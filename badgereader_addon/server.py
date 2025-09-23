import datetime
import gspread
import json
import logging
import openpyxl
import os

from aiohttp import web
from datetime import datetime, timedelta
from homeassistant_api import Client
from oauth2client.service_account import ServiceAccountCredentials

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class Config:
    def __init__(self, config_path='/data/options.json'):
        # Default values
        self.notification_domain = "notify"
        self.notification_service = "gmail_solalindenstein"
        self.notification_emails = []
        self.badges = []
        self.people = []
        self.swipe_debounce_minutes = 1
        self.swipe_time_buffer_minutes = 3
        self.storage_backend = "google_sheets"
        self.storage_file_path = "/share/badge-reader-mount/swipe_log.jsonl"
        self.storage_sheets_dir = "/share/badge-reader-mount"
        self.google_spreadsheet_url = 'https://docs.google.com/spreadsheets/d/1qZ3-8Q3z4Nn3q_V3RArP3p8G_sWn3j8aC5H6j2k_4zE/edit#gid=0' # Dummy URL, replace with your actual URL
        self.google_worksheet_name = 'Data'
        self.version = "unknown"
        self.people_map = {}
        self.badge_map = {}
        self.badge_to_person_map = {}

        self.load_from_file(config_path)

    def load_from_file(self, config_path):
        try:
            with open(config_path, 'r') as f:
                options = json.load(f)
                self.notification_domain = options.get('notification_domain', self.notification_domain)
                self.notification_service = options.get('notification_service', self.notification_service)
                self.notification_emails = options.get('notification_emails', self.notification_emails)
                self.badges = options.get('badges', self.badges)
                self.people = options.get('people', self.people)
                self.swipe_debounce_minutes = options.get('swipe_debounce_minutes', self.swipe_debounce_minutes)
                self.swipe_time_buffer_minutes = options.get('swipe_time_buffer_minutes', self.swipe_time_buffer_minutes)
                self.storage_backend = options.get('storage_backend', self.storage_backend)
                self.storage_file_path = options.get('storage_file_path', self.storage_file_path)
                self.storage_sheets_dir = options.get('storage_sheets_dir', self.storage_sheets_dir)
                self.google_spreadsheet_url = options.get('google_spreadsheet_url', self.google_spreadsheet_url)
                self.google_worksheet_name = options.get('google_worksheet_name', self.google_worksheet_name)
                # self.version = config.get('version', self.version) # Version is not in options.json
        except FileNotFoundError:
            logging.error(f"{config_path} not found. Please ensure it exists in the add-on's data directory.")
        except json.JSONDecodeError as e:
            logging.error(f"Error parsing {config_path}: {e}")

        # Create lookups for easier access
        self.people_map = {person['id']: person for person in self.people}
        self.badge_map = {badge['uid'].lower().strip(): badge['peopleID'] for badge in self.badges}
        self.badge_to_person_map = {uid: self.people_map.get(pid) for uid, pid in self.badge_map.items()}


# --- Storage Classes ---
class Storage:
    def log_swipe(self, timestamp, badge_id, action_json):
        raise NotImplementedError

    def check(self):
        raise NotImplementedError

    def read_latest_states(self):
        raise NotImplementedError

    def register_shift(self, action):
        raise NotImplementedError

class GoogleSheetStorage(Storage):
    def __init__(self, config):
        self.config = config
        self.scope = ['https://www.googleapis.com/auth/spreadsheets', "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]
        self.creds_file = '/share/google_credentials_solalindenstein_docs_user.json'
        self.sheet = self._get_sheet()

    def _get_sheet(self):
        """Authenticates with Google Sheets and returns the worksheet."""
        try:
            creds = ServiceAccountCredentials.from_json_keyfile_name(self.creds_file, self.scope)
            client = gspread.authorize(creds)
            spreadsheet = client.open_by_url(self.config.google_spreadsheet_url)
            sheet = spreadsheet.worksheet(self.config.google_worksheet_name)
            return sheet
        except Exception as e:
            logging.error(f"Error accessing Google Sheet: {e}", exc_info=True)
            return None

    def log_swipe(self, timestamp, badge_id, action_json):
        if self.sheet:
            try:
                row = [timestamp, badge_id, action_json]
                self.sheet.insert_row(row, 3)
                logging.info(f"Logged to Google Sheet: {row}")
            except Exception as e:
                logging.error(f"Error logging to Google Sheet: {e}", exc_info=True)

    def check(self):
        logging.info("Checking Google Sheet access...")
        if self.sheet:
            try:
                cell_c1 = self.sheet.cell(1, 3).value # C1
                logging.info(f"Successfully accessed Google Sheet. Cell C1 contains: '{cell_c1}'")

                headers = self.sheet.row_values(2) # Assuming headers are in row 2
                expected_headers = ['Timestamp', 'BadgeID', 'Action']
                if headers[:len(expected_headers)] == expected_headers:
                    logging.info(f"Google Sheet headers are correct: {headers}")
                else:
                    logging.warning(f"Google Sheet headers are not as expected. "
                                    f"Expected: {expected_headers}, Found: {headers}")
            except Exception as e:
                logging.error(f"Error reading from Google Sheet: {e}")
        else:
            logging.error("Could not access Google Sheet.")

    def read_latest_states(self):
        logging.info("Reading latest states from Google Sheet...")
        if not self.sheet:
            return {}
        try:
            all_records = self.sheet.get_all_values() # Gets all data, headers included
            swipe_records = all_records[2:] # Skip header rows
            swipe_records.reverse()

            latest_states = {}
            all_badge_uids = self.config.badge_map.keys()

            for record in swipe_records:
                if len(latest_states) == len(all_badge_uids):
                    break # Found the latest state for all known badges

                timestamp_str, badge_id, action_str = record[0], record[1], record[2]
                badge_id = badge_id.strip().lower()

                if badge_id in all_badge_uids and badge_id not in latest_states:
                    action = json.loads(action_str)
                    latest_states[badge_id] = {
                        'state': action.get('new_state'),
                        'timestamp': int(timestamp_str)
                    }
            logging.info(f"Found {len(latest_states)} latest states in Google Sheet.")
            return latest_states
        except Exception as e:
            logging.error(f"Error reading latest states from Google Sheet: {e}", exc_info=True)
            return {}

    def register_shift(self, action):
        pass

class FileStorage(Storage):
    def __init__(self, config):
        self.config = config
        self.file_path = self.config.storage_file_path
        self.sheets_dir = self.config.storage_sheets_dir

    def log_swipe(self, timestamp, badge_id, action_json):
        try:
            with open(self.file_path, 'a') as f:
                log_entry = {
                    'timestamp': timestamp,
                    'badge_id': badge_id,
                    'action': json.loads(action_json)
                }
                f.write(json.dumps(log_entry) + '\n')
            logging.info(f"Logged to file: {log_entry}")
        except Exception as e:
            logging.error(f"Error logging to file: {e}", exc_info=True)
    
    def check(self):
        logging.info(f"Checking file storage access at {self.file_path}...")
        try:
            with open(self.file_path, 'a') as f:
                pass
            logging.info("File storage is accessible.")
        except Exception as e:
            logging.error(f"Error accessing file storage: {e}", exc_info=True)
    
    def read_latest_states(self):
        logging.info("Reading latest states from file...")
        latest_states = {}
        try:
            if not os.path.exists(self.file_path):
                logging.warning(f"Storage file not found: {self.file_path}")
                return {}

            with open(self.file_path, 'r') as f:
                for line in f:
                    log_entry = json.loads(line)
                    badge_id = log_entry['badge_id'].strip().lower()
                    if badge_id in self.config.badge_map:
                        latest_states[badge_id] = {
                            'state': log_entry['action'].get('new_state'),
                            'timestamp': log_entry['timestamp']
                        }
            logging.info(f"Found {len(latest_states)} latest states in file.")
            return latest_states
        except Exception as e:
            logging.error(f"Error reading latest states from file: {e}", exc_info=True)
            return {}

    def register_shift(self, action):
        person_name = self.config.people_map.get(action['person_id'], {}).get('name')
        full_filename = f"{self.sheets_dir}/monthly_data_{person_name}.xlsx"

        # create file if doesnt exist yet
        if not os.path.exists(full_filename):
            logging.warning(f"Sheets not found, dir: {self.sheets_dir}")
            os.listdir(self.sheets_dir)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data"
            ws.append(["Badge Reader Data"])
            ws.append(["Person", "Shift Start", "Shift End", "Shift Duration (min)"])
            wb.save(filename = full_filename)

        wb = openpyxl.load_workbook(filename = full_filename)
        ws = wb.active
        ws.append(action)
        wb.save(filename = full_filename)

# --- Global constants and objects ---
PORT = 8199
ACCESS_KEY = "SolalindensteinBadgeReaderSecret"
HA_URL = "http://supervisor/core/api/"
HA_TOKEN = os.environ.get("SUPERVISOR_TOKEN")

# Load configuration
config = Config()

# Initialize storage backend
if config.storage_backend == 'google_sheets':
    storage = GoogleSheetStorage(config)
elif config.storage_backend == 'file':
    storage = FileStorage(config)
else:
    raise ValueError(f"Invalid storage backend: {config.storage_backend}")

last_swipe_times = {}  # Stores the last swipe time for each UID
shift_state = {} # uid -> 'in'/'out'
shift_start_times = {} # uid -> datetime

# --- App Initialization ---
def initialize_states():
    """Initializes shift states from storage at startup."""
    logging.info("Initializing states from storage...")
    initial_states = storage.read_latest_states()

    # Set default 'out' state for all badges
    for uid in config.badge_map.keys():
        shift_state[uid] = 'out'

    # Restore latest known states
    for badge_uid, data in initial_states.items():
        if data and 'state' in data and 'timestamp' in data:
            shift_state[badge_uid] = data['state']
            person = config.badge_to_person_map.get(badge_uid)
            person_name = person['name'] if person else 'Unknown'

            if data['state'] == 'in':
                start_time = datetime.fromtimestamp(data['timestamp'])
                shift_start_times[badge_uid] = start_time
                logging.info(f"Restored state for {person_name} ({badge_uid}): IN, started at {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
            else:
                logging.info(f"Restored state for {person_name} ({badge_uid}): OUT")
        else:
             logging.warning(f"Could not restore state for badge {badge_uid}, data is incomplete: {data}")

logging.info(f"Loaded {len(config.badges)} badges and {len(config.people)} people.")
logging.info(f"Notification emails: {config.notification_emails}")
logging.info(f"Swipe debounce time set to {config.swipe_debounce_minutes} minute(s).")
logging.info(f"Swipe time buffer set to {config.swipe_time_buffer_minutes} minute(s).")


async def send_notification(title, message, person=None):
    """Sends a notification using the Home Assistant notify service."""
    if not HA_TOKEN:
        logging.warning("SUPERVISOR_TOKEN not found, cannot send notification.")
        return

    if person and person.get('email'):
        primary_recipients = [person['email']]
        cc_recipients = [email for email in config.notification_emails if email != person['email']]
    else:
        primary_recipients = config.notification_emails
        cc_recipients = []

    if not primary_recipients:
        logging.warning("No recipients found for notification.")
        return

    service_data = {
        'title': title,
        'message': message,
        'target': primary_recipients
    }

    if cc_recipients:
        service_data['data'] = {'cc': cc_recipients}

    logging.info(f"Attempting to send notification to {primary_recipients} (CC: {cc_recipients}). "
                 f"Domain: '{config.notification_domain}', Service: '{config.notification_service}', Data: {service_data}")

    try:
        async with Client(HA_URL, HA_TOKEN, use_async=True) as client:
            await client.async_trigger_service(config.notification_domain, config.notification_service, **service_data)
        logging.info(f"Successfully sent notification with title: '{title}' to {primary_recipients}")
    except Exception as e:
        logging.error(f"Error sending notification to {primary_recipients}: {e}", exc_info=True)

async def process_card_swipe(card_uid, data):
    """Processes a card swipe after the UID has been extracted."""
    sanitized_card_uid = str(card_uid).strip().lower()

    if sanitized_card_uid in config.badge_map:
        people_id = config.badge_map[sanitized_card_uid]
        person = config.people_map.get(people_id)

        if not person:
            logging.error(f"Badge with UID {sanitized_card_uid} is mapped to a non-existent person with ID {people_id}")
            return web.HTTPInternalServerError(text="Internal server error - bad configuration")

        person_name = person['name']

        # Check for multi-swipes
        now = datetime.now()
        last_swipe = last_swipe_times.get(sanitized_card_uid)

        if last_swipe:
            time_since_last_swipe = now - last_swipe
            if time_since_last_swipe < timedelta(minutes=config.swipe_debounce_minutes):
                logging.warning(
                    f"Ignoring duplicate swipe for {person_name} ({sanitized_card_uid}). "
                    f"Last swipe was {time_since_last_swipe.total_seconds():.1f} seconds ago. "
                    f"Debounce is set to {config.swipe_debounce_minutes} minute(s)."
                )
                return web.Response(text=f"Duplicate swipe for {person_name}. Please wait.")

        last_swipe_times[sanitized_card_uid] = now
        unix_timestamp = int(now.timestamp())
        logging.info(f"Recognized card for: {person_name}")

        current_state = shift_state.get(sanitized_card_uid, 'out')
        buffer = timedelta(minutes=config.swipe_time_buffer_minutes)

        if current_state == 'out':
            new_state = 'in'
            shift_state[sanitized_card_uid] = new_state
            shift_start_times[sanitized_card_uid] = now
            
            effective_time = (now - buffer).replace(second=0, microsecond=0)
            action = {
                'person_id': people_id,
                'new_state': new_state,
                'time_effective': int(effective_time.timestamp())
            }
            storage.log_swipe(unix_timestamp, sanitized_card_uid, json.dumps(action))

            await send_notification(
                title=f"{person_name} - Schicht gestartet",
                message=f"Hallo {person_name}, deine Schicht hat gerade begonnen. Einen schönen und produktiven Tag!",
                person=person
            )
            return web.Response(text=f"Welcome, {person_name}. Your shift has started.")
        else: # current_state == 'in'
            new_state = 'out'
            shift_state[sanitized_card_uid] = new_state
            start_time = shift_start_times.pop(sanitized_card_uid, now)
            duration = now - start_time
            
            total_seconds = duration.total_seconds()
            hours = int(total_seconds // 3600)
            minutes = int((total_seconds % 3600) // 60)
            duration_str = f"{hours} Stunden und {minutes} Minuten"

            effective_time = (now + buffer).replace(second=0, microsecond=0)
            action = {
                'person_id': people_id,
                'new_state': new_state,
                'time_effective': int(effective_time.timestamp()),
                'shift_start_timestamp': start_time.timestamp(),
                'shift_end_timestamp': effective_time.timestamp(),
                'shift_duration_min': minutes
            }
            storage.log_swipe(unix_timestamp, sanitized_card_uid, json.dumps(action))
            storage.register_shift(action)

            await send_notification(
                title=f"{person_name} - Schicht beendet",
                message=f"Hallo {person_name}, deine Schicht ist nun zu Ende. Deine heutige Arbeitszeit betrug {duration_str}. Wir wünschen dir einen schönen Feierabend!",
                person=person
            )
            return web.Response(text=f"Goodbye, {person_name}. Your shift has ended.")

    else:
        logging.warning(f"Unrecognized card: {card_uid}. Full request data: {data}")
        await send_notification(
            title="Unbekannter Badge gescannt",
            message=f"Hallo, gerade wurde ein unbekannter Badge mit der UID {card_uid} gescannt."
        )
        raise web.HTTPUnauthorized(text="Unrecognized card")


async def handle_post(request):
    try:
        data = await request.post()
        access_key = request.query.get('accessKey') or data.get('accessKey')

        if access_key != ACCESS_KEY:
            logging.warning(f"Unauthorized request from {request.remote}. URL query: '{request.query_string}'. POST data: {data}")
            raise web.HTTPUnauthorized(text="Unauthorized")

        card_uid = data.get("UID")
        if not card_uid:
            logging.warning(f"Received data with no 'UID' field: {data}")
            raise web.HTTPBadRequest(text="Missing 'UID' in payload")

        logging.info(f"Extracted card UID: {card_uid}")
        return await process_card_swipe(card_uid, data)

    except web.HTTPException as e:
        raise e
    except Exception as e:
        logging.error(f"Error processing POST request: {e}", exc_info=True)
        raise web.HTTPInternalServerError(text="Internal server error")

async def handle_get(request):
    html_response = '''
    <html>
        <head><title>Badge Reader Addon Server</title></head>
        <body>
            <h1>Badge Reader Addon HTTP Server</h1>
            <p>This server is running and listening for POST requests for badge data.</p>
        </body>
    </html>
    '''
    logging.info(f"Served GET request to {request.path} from {request.remote}")
    return web.Response(text=html_response, content_type='text/html')

app = web.Application()
app.add_routes([
    web.post('/', handle_post),
    web.get('/', handle_get),
])

if __name__ == "__main__":
    if not HA_TOKEN:
        logging.warning("SUPERVISOR_TOKEN environment variable not set. Home Assistant integration will be disabled.")

    storage.check()
    initialize_states()

    logging.info(f"Hello from Badge Reader server, version {config.version}")
    logging.info(f"Starting HTTP server for badge reader on port {PORT}...")
    logging.info(f"Server listening on 0.0.0.0:{PORT}")
    with open("/share/badge-reader-mount/testfile.txt") as f:
        print f.read()
    logging.info(f"Badge messages should be sent to http://<ADDON_IP_ADDRESS>:{PORT}/?accessKey={ACCESS_KEY}")
    web.run_app(app, host='0.0.0.0', port=PORT)
