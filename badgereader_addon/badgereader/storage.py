import json
import logging
import os
import openpyxl
import locale
from libnfs import NFS
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl.styles import Font

# --- Storage Classes ---
class Storage:
    def log_swipe(self, timestamp, badge_id, action_json):
        raise NotImplementedError

    def check(self):
        raise NotImplementedError

    def read_latest_states(self):
        raise NotImplementedError

    def register_shift(self, action):
        raise NotImplementedError


class NFSStorage(Storage):
    def __init__(self, config):
        self.config = config
        nfs_url = f"nfs://{self.config.nfs_server_address}{self.config.nfs_share_path}"
        self.nfs = NFS(nfs_url)
        self.log_file_path = "swipe_log.jsonl"
        # Set locale for German month names
        try:
            locale.setlocale(locale.LC_TIME, 'de_DE.UTF-8')
        except locale.Error:
            logging.warning("Could not set locale to de_DE.UTF-8. Month names may not be in German.")


    def log_swipe(self, timestamp, badge_id, action_json):
        f = None
        try:
            log_entry = {
                "timestamp": timestamp,
                "badge_id": badge_id,
                "action": json.loads(action_json),
            }
            log_line = json.dumps(log_entry) + "\n"
            f = self.nfs.open(self.log_file_path, "a")
            f.write(bytearray(log_line, 'utf-8'))
            logging.info(f"Logged to NFS: {log_entry}")
        except Exception as e:
            logging.error(f"Error logging to NFS: {e}", exc_info=True)
        finally:
            if f:
                f.close()

    def check(self):
        logging.info(f"Checking NFS storage access at {self.config.nfs_server_address}:{self.config.nfs_share_path}...")
        try:
            self.nfs.listdir(".")
            logging.info("NFS storage is accessible.")
        except Exception as e:
            logging.error(f"Error accessing NFS storage: {e}", exc_info=True)

    def read_latest_states(self):
        logging.info("Reading latest states from NFS...")
        latest_states = {}
        f = None
        try:
            if not self.nfs.isfile(self.log_file_path):
                logging.warning(f"Storage file not found on NFS: {self.log_file_path}")
                return {}

            f = self.nfs.open(self.log_file_path, "r")
            content = f.read()
            for line in content.splitlines():
                log_entry = json.loads(line)
                badge_id = log_entry["badge_id"].strip().lower()
                if badge_id in self.config.badge_map:
                    latest_states[badge_id] = {
                        "state": log_entry["action"].get("new_state"),
                        "timestamp": log_entry["timestamp"],
                    }
            logging.info(f"Found {len(latest_states)} latest states in NFS.")
            return latest_states
        except Exception as e:
            logging.error(f"Error reading latest states from NFS: {e}", exc_info=True)
            return {}
        finally:
            if f:
                f.close()

    def _get_sheet_filename(self, person_name, date_obj):
        month_str = date_obj.strftime('%B')
        year = date_obj.year
        return f"Arbeitszeit {person_name} {month_str} {year}.xlsx"

    def _create_new_sheet(self, filename, initial_balance=0):
        f = None
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Daten"
            ws['A1'] = "Ausweisleser Daten"

            ws['A2'] = "Zeitguthaben aus dem Vormonat (min)"
            ws['B2'] = initial_balance

            ws['A3'] = "Soll-Stunden pro Schicht (min)"
            ws['B3'] = 300  # 5 hours * 60 minutes

            ws['A4'] = "Aktueller Laufender Saldo (min)"
            ws['B4'] = f'=SUM($B$2, G6:G500)' # Formula to auto-update the running balance

            # Headers for the data table
            headers = ["Tag", "Ereignis", "Schichtbeginn", "Schichtende", "Schichtdauer (HH:MM)", "Tages-Soll (min)", "Tages-Nettoveränderung (min)", "Laufender Saldo (min)"]
            ws.append(headers)
            
            # Make header bold
            bold_font = Font(bold=True)
            for cell in ws[5]: # Headers are on row 5
                cell.font = bold_font

            # Autosize columns
            for column_cells in ws.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            f = self.nfs.open(filename, "wb")
            f.write(bytearray(buffer.read()))
            logging.info(f"Successfully created new sheet: {filename}")
            return True
        except Exception as e:
            logging.error(f"Error creating new sheet {filename}: {e}", exc_info=True)
            return False
        finally:
            if f:
                f.close()

    def _ensure_sheet_exists(self, person_name, now):
        current_filename = self._get_sheet_filename(person_name, now)

        if self.nfs.isfile(current_filename):
            return current_filename

        logging.info(f"Sheet for {person_name} for {now.strftime('%B')} {now.year} not found. Looking for previous month's sheet.")

        first_day_of_current_month = now.replace(day=1)
        last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)
        prev_month_filename = self._get_sheet_filename(person_name, last_day_of_previous_month)

        initial_balance = 0
        if self.nfs.isfile(prev_month_filename):
            logging.info(f"Found previous month's sheet: {prev_month_filename}. Reading end of month balance.")
            f = None
            try:
                f = self.nfs.open(prev_month_filename, "rb")
                file_content = f.read()
                workbook_stream = BytesIO(file_content)
                # Use data_only=True to read the calculated value of a formula
                wb_data_only = openpyxl.load_workbook(workbook_stream, data_only=True)
                ws_data_only = wb_data_only.active
                balance_value = ws_data_only['B4'].value
                if balance_value is not None:
                    initial_balance = float(balance_value)
                logging.info(f"Balance from previous month is {initial_balance}")

            except Exception as e:
                logging.error(f"Error reading balance from {prev_month_filename}: {e}", exc_info=True)
            finally:
                if f:
                    f.close()

        if self._create_new_sheet(current_filename, initial_balance):
            return current_filename

        return None  # Indicate failure

    def _append_to_sheet(self, filename, action):
        f_read = None
        f_write = None
        try:
            f_read = self.nfs.open(filename, "rb")
            file_content = f_read.read()
            workbook_stream = BytesIO(file_content)
            wb = openpyxl.load_workbook(workbook_stream)

            ws = wb.active

            new_row_num = ws.max_row + 1

            shift_start_ts = action.get('shift_start_timestamp')
            shift_end_ts = action.get('shift_end_timestamp')
            
            if not shift_start_ts or not shift_end_ts:
                logging.error(f"Action is missing timestamp information: {action}")
                return None
            
            start_dt = datetime.fromtimestamp(shift_start_ts)
            end_dt = datetime.fromtimestamp(shift_end_ts)
            
            day_str = start_dt.strftime('%Y-%m-%d')
            start_time_str = start_dt.strftime('%H:%M')
            end_time_str = end_dt.strftime('%H:%M')
            event_str = "Arbeitsschicht"

            shift_duration_minutes = action.get('shift_duration_min', 0)
            
            duration_hours = shift_duration_minutes // 60
            duration_mins_rem = shift_duration_minutes % 60
            duration_str = f"{duration_hours:02d}:{duration_mins_rem:02d}"

            target_for_day = ws['B3'].value
            net_change = shift_duration_minutes - target_for_day
            
            # Formula for the running balance in column H
            running_balance_formula = f"=SUM($B$2,G$6:G{new_row_num})"

            row_data = [
                day_str,
                event_str,
                start_time_str,
                end_time_str,
                duration_str,
                target_for_day,
                net_change,
                running_balance_formula
            ]
            ws.append(row_data)

            # Autosize columns
            for column_cells in ws.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            f_write = self.nfs.open(filename, "wb")
            f_write.write(bytearray(buffer.read()))
            logging.info(f"Successfully registered shift in {filename}")
            return True # Return success
        except Exception as e:
            logging.error(f"Error appending to sheet {filename}: {e}", exc_info=True)
            return None
        finally:
            if f_read:
                f_read.close()
            if f_write:
                f_write.close()

    def register_shift(self, action):
        person_name = action.get("name")
        if not person_name:
            person_id = action.get("person_id")
            person_info = self.config.people_map.get(person_id)
            if person_info:
                person_name = person_info.get("name")

        if not person_name:
            logging.error(f"Could not determine person name from action: {action}")
            return None

        now = datetime.now()
        sheet_filename = self._ensure_sheet_exists(person_name, now)

        if sheet_filename:
            return self._append_to_sheet(sheet_filename, action)
        else:
            logging.error(f"Failed to ensure sheet exists for {person_name}. Cannot register shift.")
            return None
