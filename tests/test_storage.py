
import pytest
from unittest.mock import MagicMock, patch, call, ANY
from datetime import datetime
from io import BytesIO
import openpyxl

# Mock NFS before it's imported by the storage module
with patch.dict("sys.modules", {"libnfs": MagicMock()}):
    from badgereader_addon.badgereader.storage import NFSStorage

@pytest.fixture
def mock_config():
    """Fixture to create a mock config object."""
    config = MagicMock()
    config.nfs_server_address = "mock_server"
    config.nfs_share_path = "/mock_share"
    config.people_map = {"person1": {"name": "Fin"}}
    return config

@pytest.fixture
def mock_nfs_instance():
    """Fixture for a mocked NFS instance."""
    return MagicMock()

@pytest.fixture
def nfs_storage(mock_config, mock_nfs_instance):
    """Fixture for an NFSStorage instance with a mocked NFS client."""
    with patch("badgereader_addon.badgereader.storage.NFS", return_value=mock_nfs_instance):
        storage = NFSStorage(config=mock_config)
        return storage

def test_get_sheet_filename(nfs_storage):
    """Test the generation of the Excel sheet filename."""
    date_obj = datetime(2024, 7, 10)
    filename = nfs_storage._get_sheet_filename("Fin", date_obj)
    assert filename == "Arbeitszeit Fin July 2024.xlsx"

def test_create_new_sheet(nfs_storage, mock_nfs_instance):
    """Test the creation of a new timesheet from scratch."""
    filename = "test_sheet.xlsx"

    with patch("openpyxl.Workbook") as mock_workbook_class:
        mock_wb = MagicMock()
        mock_ws = MagicMock()
        mock_wb.active = mock_ws
        mock_workbook_class.return_value = mock_wb

        # Correctly mock the worksheet's __setitem__ to capture the writes
        cell_values = {}
        def set_item_side_effect(key, value):
            cell_values[key] = value
        mock_ws.__setitem__.side_effect = set_item_side_effect

        # Mock the context manager for the file write
        mock_file_write = MagicMock()
        mock_nfs_instance.open.return_value.__enter__.return_value = mock_file_write

        nfs_storage._create_new_sheet(filename, initial_balance=50)

        # Verify worksheet setup by checking the captured cell values
        assert cell_values['A2'] == "Time Balance from Previous Month (min)"
        assert cell_values['B2'] == 50
        assert cell_values['A4'] == "Current Running Balance (min)"

        # Verify headers were added via append
        mock_ws.append.assert_called_once()
        mock_nfs_instance.open.assert_called_with(filename, "wb")
        mock_wb.save.assert_called_once()


def test_ensure_sheet_exists_already_exists(nfs_storage, mock_nfs_instance):
    """Test that if the sheet for the current month exists, it is returned."""
    now = datetime(2024, 7, 10)
    person_name = "Fin"
    current_filename = nfs_storage._get_sheet_filename(person_name, now)
    mock_nfs_instance.exists.return_value = True

    filename = nfs_storage._ensure_sheet_exists(person_name, now)

    mock_nfs_instance.exists.assert_called_once_with(current_filename)
    assert filename == current_filename

def test_ensure_sheet_exists_copies_previous(nfs_storage, mock_nfs_instance):
    """Test that a new sheet is created by copying the previous month's sheet."""
    now = datetime(2024, 8, 1)
    person_name = "Fin"
    current_filename = nfs_storage._get_sheet_filename(person_name, now)
    prev_filename = nfs_storage._get_sheet_filename(person_name, now.replace(day=1) - timedelta(days=1))

    mock_nfs_instance.exists.side_effect = [False, True]

    wb_data_only = openpyxl.Workbook()
    ws_data_only = wb_data_only.active
    ws_data_only['B4'] = 120.0  # Make it a float like openpyxl would read
    buffer = BytesIO()
    wb_data_only.save(buffer)
    buffer.seek(0)

    mock_file = MagicMock()
    mock_file.read.return_value = buffer.getvalue()
    mock_nfs_instance.open.return_value.__enter__.return_value = mock_file

    with patch.object(nfs_storage, '_create_new_sheet') as mock_create:
        nfs_storage._ensure_sheet_exists(person_name, now)
        mock_nfs_instance.exists.assert_has_calls([call(current_filename), call(prev_filename)])
        mock_nfs_instance.open.assert_called_with(prev_filename, "rb")
        mock_create.assert_called_once_with(current_filename, 120.0)

def test_append_to_sheet(nfs_storage, mock_nfs_instance):
    """Test that a new shift is correctly appended and the balance is updated."""
    filename = "test_sheet.xlsx"
    action = {
        'name': 'Fin',
        'shift_start': '2024-07-10T09:00:00',
        'shift_end': '2024-07-10T14:30:00',
        'duration_minutes': 330
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B2"] = 50
    ws["B3"] = 300
    ws.append(["H1", "H2", "H3", "H4", "H5", "H6", "H7"])
    ws.append(["D1", "D2", "D3", "D4", "D5", "D6", 100]) # Last row data

    input_buffer = BytesIO()
    wb.save(input_buffer)
    input_buffer.seek(0)

    # Mock the read operation
    mock_read_file = MagicMock()
    mock_read_file.read.return_value = input_buffer.getvalue()
    # Mock the write operation
    mock_write_buffer = BytesIO()

    def open_side_effect(path, mode):
        if mode == "rb":
            return MagicMock(__enter__=MagicMock(return_value=mock_read_file))
        elif mode == "wb":
            # This allows us to inspect the saved workbook later if needed
            return MagicMock(__enter__=MagicMock(return_value=BytesIO()))
    mock_nfs_instance.open.side_effect = open_side_effect

    new_balance = nfs_storage._append_to_sheet(filename, action)

    assert new_balance == 180 # 100 (from last row) + 330 (shift) - 300 (target)
    assert mock_nfs_instance.open.call_count == 2


def test_register_shift_orchestrates_correctly(nfs_storage):
    """Test that register_shift calls the helper functions correctly."""
    action = {'person_id': 'person1', 'name': 'Fin', 'duration_minutes': 315}
    sheet_filename = "Arbeitszeit Fin July 2024.xlsx"
    expected_balance = 100

    with patch.object(nfs_storage, '_ensure_sheet_exists', return_value=sheet_filename) as mock_ensure, \
         patch.object(nfs_storage, '_append_to_sheet', return_value=expected_balance) as mock_append, \
         patch("badgereader_addon.badgereader.storage.datetime") as mock_dt:

        mock_dt.now.return_value = datetime(2024, 7, 10)
        balance = nfs_storage.register_shift(action)

        mock_ensure.assert_called_once_with('Fin', mock_dt.now.return_value)
        mock_append.assert_called_once_with(sheet_filename, action)
        assert balance == expected_balance
